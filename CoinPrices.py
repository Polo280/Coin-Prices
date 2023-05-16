import openpyxl
from bs4 import BeautifulSoup
import requests

def updateValues():
    path = "C:/Users/jorgl/OneDrive/Escritorio/Guide.xlsx"

    try:
        cryptos = ["bitcoin", "ethereum", "bnb", "cardano", "solana", "polkadot-new", "monero",
                   "basic-attention-token", "harmony"]

        cryptoExcel = ["Bitcoin (BTC)", "Ethereum (ETH)", "Binance Coin (BNB)", "Cardano (ADA)", "Solana (SOL)",
                       "Polkadot (DOT)", "Monero (XMR)", "Brave Bat", "Harmony ONE"]

        agent = {"User-Agent": "Brave"}
        priceList = []

        for item in cryptos:
            site = "https://coinmarketcap.com/currencies/{}/".format(item)
            request = requests.get(site, headers=agent)

            if request.status_code == 200:
                b = BeautifulSoup(request.content, "html.parser")
                price = b.find('div', attrs={"class": "priceValue"})  # Search for it in raw html
                if price is not None:
                    normPrice = float(price.text.removeprefix("$").replace(",", ""))  # Convert into float
                    priceList.append(normPrice)
                else:
                    # With the newest update of CMC, some cryptos dont have div class "priceValue", instead the price is inside span class sc-8755d3ba-0 hQJDQt
                    price = b.find('span', attrs={"class": "sc-8755d3ba-0 hQJDQt"})
                    normPrice = float(price.text.removeprefix("\xa0$").replace(",", ""))  # Convert into float
                    priceList.append(normPrice)
            else:
                print("Request failed with code {}".format(request.status_code))

        wb = openpyxl.load_workbook(path, read_only=False, keep_vba=False)
        sheet1 = wb.active

        index = 0
        for column in sheet1.iter_cols(min_col=1, max_col=1, min_row=2, max_row=len(cryptos)+1):
            for cell in column:
                cell.value = cryptoExcel[index]
                index += 1

        index = 0
        for column in sheet1.iter_cols(min_col= 2, max_col= 2, min_row= 2, max_row=len(cryptos)+1):
            for cell in column:
                cell.value = priceList[index]
                index += 1
        wb.save(path)
        print("Excel file saved with success\n")

    except Exception as ex:
        print("Program failed due to exception")
        print(ex)


if __name__ == '__main__':
    try:
        updateValues()
    except KeyboardInterrupt:
        print("Program terminated by user")